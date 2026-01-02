"""Attachment processing service for extracting content from issue attachments."""
import os
import io
import logging
import tempfile
import mimetypes
import zipfile
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import requests
import urllib3

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logger = logging.getLogger(__name__)

# Supported file types for text extraction
TEXT_EXTENSIONS = {
    '.txt', '.csv', '.json', '.xml', '.md', '.log', '.yaml', '.yml',
    '.ini', '.cfg', '.conf', '.html', '.htm'
}
CODE_EXTENSIONS = {
    '.py', '.java', '.js', '.ts', '.c', '.cpp', '.h', '.cs', '.sql', 
    '.sh', '.bat', '.ps1', '.rb', '.go', '.rs', '.swift', '.kt', 
    '.scala', '.r', '.m', '.php', '.hpp', '.css', '.jsx', '.tsx',
    '.vue', '.svelte', '.lua', '.pl', '.pm', '.tcl', '.awk', '.sed',
    '.makefile', '.cmake', '.gradle', '.groovy', '.dart', '.elm',
    '.erl', '.ex', '.exs', '.hs', '.ml', '.fs', '.clj', '.lisp'
}
DOCUMENT_EXTENSIONS = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.odt', '.ods', '.odp'}
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif', '.webp'}
ARCHIVE_EXTENSIONS = {'.zip', '.tar', '.gz', '.tgz', '.tar.gz', '.7z', '.rar'}

# Maximum file size for processing (10MB)
MAX_FILE_SIZE = 10 * 1024 * 1024


class AttachmentService:
    """Service for downloading and extracting content from issue attachments."""
    
    def __init__(self, temp_dir: Optional[str] = None, enable_ocr: bool = False):
        """
        Initialize the attachment service.
        
        Args:
            temp_dir: Directory for temporary file storage
            enable_ocr: Enable OCR for image-based documents (requires tesseract)
        """
        self.temp_dir = temp_dir or tempfile.gettempdir()
        self.enable_ocr = enable_ocr
        
        # Lazy-load optional dependencies
        self._pypdf = None
        self._docx = None
        self._openpyxl = None
        self._pptx = None
        self._pytesseract = None
        self._pillow = None
        
        logger.info(f"Attachment service initialized (temp_dir: {self.temp_dir}, ocr: {enable_ocr})")
    
    def _get_pypdf(self):
        """Lazy load PyPDF2."""
        if self._pypdf is None:
            try:
                import pypdf
                self._pypdf = pypdf
            except ImportError:
                logger.warning("pypdf not installed. PDF extraction disabled.")
        return self._pypdf
    
    def _get_docx(self):
        """Lazy load python-docx."""
        if self._docx is None:
            try:
                import docx
                self._docx = docx
            except ImportError:
                logger.warning("python-docx not installed. DOCX extraction disabled.")
        return self._docx
    
    def _get_openpyxl(self):
        """Lazy load openpyxl."""
        if self._openpyxl is None:
            try:
                import openpyxl
                self._openpyxl = openpyxl
            except ImportError:
                logger.warning("openpyxl not installed. XLSX extraction disabled.")
        return self._openpyxl
    
    def _get_pillow(self):
        """Lazy load PIL."""
        if self._pillow is None:
            try:
                from PIL import Image
                self._pillow = Image
            except ImportError:
                logger.warning("Pillow not installed. Image processing disabled.")
        return self._pillow
    
    def download_attachment(self, url: str, filename: str, 
                           auth_header: Optional[Dict[str, str]] = None) -> Optional[str]:
        """
        Download an attachment from URL to temp directory.
        
        Args:
            url: Attachment URL
            filename: Original filename
            auth_header: Optional authentication headers
            
        Returns:
            Local file path or None on failure
        """
        try:
            headers = auth_header or {}
            response = requests.get(url, headers=headers, verify=False, stream=True, timeout=30)
            response.raise_for_status()
            
            # Check file size
            content_length = int(response.headers.get('content-length', 0))
            if content_length > MAX_FILE_SIZE:
                logger.warning(f"Attachment too large: {filename} ({content_length} bytes)")
                return None
            
            # Save to temp file
            safe_filename = "".join(c for c in filename if c.isalnum() or c in '._-')
            file_path = os.path.join(self.temp_dir, f"attachment_{safe_filename}")
            
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            logger.info(f"Downloaded attachment: {filename} -> {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Failed to download attachment {filename}: {e}")
            return None
    
    def extract_text_from_file(self, file_path: str) -> Tuple[str, str]:
        """
        Extract text content from a file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Tuple of (extracted_text, file_type)
        """
        ext = Path(file_path).suffix.lower()
        
        try:
            if ext in TEXT_EXTENSIONS or ext in CODE_EXTENSIONS:
                return self._extract_text_file(file_path), 'text'
            elif ext == '.pdf':
                return self._extract_pdf(file_path), 'pdf'
            elif ext in {'.doc', '.docx'}:
                return self._extract_docx(file_path), 'document'
            elif ext in {'.xls', '.xlsx'}:
                return self._extract_xlsx(file_path), 'spreadsheet'
            elif ext in {'.zip'}:
                return self._extract_zip(file_path), 'archive'
            elif ext in IMAGE_EXTENSIONS:
                return self._extract_image(file_path), 'image'
            else:
                logger.info(f"Unsupported file type: {ext}")
                return "", 'unsupported'
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {e}")
            return "", 'error'
    
    def _extract_text_file(self, file_path: str) -> str:
        """Extract text from plain text files."""
        encodings = ['utf-8', 'utf-16', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        
        # Binary fallback
        with open(file_path, 'rb') as f:
            return f.read().decode('utf-8', errors='ignore')
    
    def _extract_pdf(self, file_path: str) -> str:
        """Extract text from PDF files."""
        pypdf = self._get_pypdf()
        if not pypdf:
            return "[PDF extraction not available - install pypdf]"
        
        text_parts = []
        try:
            with open(file_path, 'rb') as f:
                reader = pypdf.PdfReader(f)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
        except Exception as e:
            logger.error(f"PDF extraction error: {e}")
            return f"[PDF extraction failed: {e}]"
        
        return "\n\n".join(text_parts)
    
    def _extract_docx(self, file_path: str) -> str:
        """Extract text from DOCX files."""
        docx_module = self._get_docx()
        if not docx_module:
            return "[DOCX extraction not available - install python-docx]"
        
        try:
            doc = docx_module.Document(file_path)
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            return "\n\n".join(paragraphs)
        except Exception as e:
            logger.error(f"DOCX extraction error: {e}")
            return f"[DOCX extraction failed: {e}]"
    
    def _extract_xlsx(self, file_path: str) -> str:
        """Extract text from XLSX files."""
        openpyxl = self._get_openpyxl()
        if not openpyxl:
            return "[XLSX extraction not available - install openpyxl]"
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text_parts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_text):
                        text_parts.append(" | ".join(row_text))
            
            return "\n".join(text_parts)
        except Exception as e:
            logger.error(f"XLSX extraction error: {e}")
            return f"[XLSX extraction failed: {e}]"
    
    def _extract_zip(self, file_path: str) -> str:
        """Extract and process text from files inside a ZIP archive."""
        text_parts = []
        files_processed = 0
        max_files = 50  # Limit files to process from archive
        max_content_per_file = 10000  # Max chars per file
        
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                # List all files in the archive
                file_list = zf.namelist()
                text_parts.append(f"=== ZIP Archive Contents ({len(file_list)} files) ===\n")
                
                for zip_info in zf.infolist():
                    if files_processed >= max_files:
                        text_parts.append(f"\n[... truncated, {len(file_list) - files_processed} more files ...]")
                        break
                    
                    # Skip directories
                    if zip_info.is_dir():
                        continue
                    
                    filename = zip_info.filename
                    ext = Path(filename).suffix.lower()
                    
                    # Only process text/code files inside the archive
                    if ext in TEXT_EXTENSIONS or ext in CODE_EXTENSIONS:
                        try:
                            with zf.open(zip_info) as f:
                                # Try to read as text
                                content = f.read()
                                try:
                                    text_content = content.decode('utf-8')
                                except UnicodeDecodeError:
                                    try:
                                        text_content = content.decode('latin-1')
                                    except:
                                        text_content = "[Binary content]"
                                
                                # Truncate if too long
                                if len(text_content) > max_content_per_file:
                                    text_content = text_content[:max_content_per_file] + "\n[... content truncated ...]"
                                
                                text_parts.append(f"\n--- {filename} ---")
                                text_parts.append(text_content)
                                files_processed += 1
                        except Exception as e:
                            text_parts.append(f"\n--- {filename} ---")
                            text_parts.append(f"[Error reading file: {e}]")
                    else:
                        # Just list non-text files
                        size_kb = zip_info.file_size / 1024
                        text_parts.append(f"  • {filename} ({size_kb:.1f} KB)")
                
                if files_processed == 0:
                    text_parts.append("\n[No text files found in archive]")
                else:
                    text_parts.append(f"\n\n=== Processed {files_processed} text files ===")
            
            return "\n".join(text_parts)
            
        except zipfile.BadZipFile:
            logger.error(f"Invalid ZIP file: {file_path}")
            return "[Invalid or corrupted ZIP file]"
        except Exception as e:
            logger.error(f"ZIP extraction error: {e}")
            return f"[ZIP extraction failed: {e}]"
    
    def _extract_image(self, file_path: str) -> str:
        """Extract information from images - OCR if available, otherwise metadata."""
        Image = self._get_pillow()
        
        # Get basic image info
        image_info = []
        filename = Path(file_path).name
        
        if Image:
            try:
                img = Image.open(file_path)
                width, height = img.size
                mode = img.mode
                format_type = img.format or "Unknown"
                image_info.append(f"Image: {filename}")
                image_info.append(f"Dimensions: {width}x{height} pixels")
                image_info.append(f"Format: {format_type}")
                image_info.append(f"Mode: {mode}")
                
                # Try OCR if enabled
                if self.enable_ocr:
                    try:
                        import pytesseract
                        text = pytesseract.image_to_string(img)
                        if text.strip():
                            image_info.append(f"\n--- OCR Extracted Text ---")
                            image_info.append(text.strip())
                    except ImportError:
                        image_info.append("[OCR not available - pytesseract not installed]")
                    except Exception as e:
                        image_info.append(f"[OCR failed: {e}]")
                else:
                    image_info.append("[Image attached - OCR disabled]")
                    
            except Exception as e:
                image_info.append(f"Image: {filename}")
                image_info.append(f"[Could not read image metadata: {e}]")
        else:
            image_info.append(f"Image: {filename}")
            image_info.append("[Image processing not available - install Pillow for metadata]")
        
        return "\n".join(image_info)
    
    def _extract_image_ocr(self, file_path: str) -> str:
        """Extract text from images using OCR (legacy method)."""
        return self._extract_image(file_path)
    
    def process_attachment(self, attachment_info: Dict[str, Any],
                          auth_header: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
        """
        Download and process a single attachment.
        
        Args:
            attachment_info: Attachment metadata (must contain 'content_url' and 'filename')
            auth_header: Optional authentication headers
            
        Returns:
            Processed attachment data with extracted text
        """
        filename = attachment_info.get('filename', 'unknown')
        url = attachment_info.get('content_url')
        
        if not url:
            return {
                'filename': filename,
                'success': False,
                'error': 'No content URL provided',
                'content': ''
            }
        
        result = {
            'filename': filename,
            'size': attachment_info.get('size', 0),
            'author': attachment_info.get('author', 'unknown'),
            'created': attachment_info.get('created', ''),
            'url': url
        }
        
        # Check if file type is supported
        ext = Path(filename).suffix.lower()
        supported_extensions = TEXT_EXTENSIONS | CODE_EXTENSIONS | DOCUMENT_EXTENSIONS | ARCHIVE_EXTENSIONS | IMAGE_EXTENSIONS
        
        if ext not in supported_extensions:
            result['success'] = False
            result['error'] = f'Unsupported file type: {ext}'
            result['content'] = ''
            result['file_type'] = 'unsupported'
            return result
        
        # Download file
        file_path = self.download_attachment(url, filename, auth_header)
        if not file_path:
            result['success'] = False
            result['error'] = 'Download failed'
            result['content'] = ''
            return result
        
        try:
            # Extract text
            content, file_type = self.extract_text_from_file(file_path)
            
            result['success'] = True
            result['content'] = content
            result['file_type'] = file_type
            result['content_length'] = len(content)
            
        finally:
            # Clean up temp file
            try:
                if file_path and os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass
        
        return result
    
    def process_attachments(self, attachments: List[Dict[str, Any]],
                           auth_header: Optional[Dict[str, str]] = None,
                           max_attachments: int = 10) -> List[Dict[str, Any]]:
        """
        Process multiple attachments.
        
        Args:
            attachments: List of attachment metadata
            auth_header: Optional authentication headers
            max_attachments: Maximum number of attachments to process
            
        Returns:
            List of processed attachment data
        """
        results = []
        
        for i, attachment in enumerate(attachments[:max_attachments]):
            logger.info(f"Processing attachment {i+1}/{min(len(attachments), max_attachments)}: {attachment.get('filename', 'unknown')}")
            result = self.process_attachment(attachment, auth_header)
            results.append(result)
        
        successful = sum(1 for r in results if r.get('success'))
        logger.info(f"Processed {successful}/{len(results)} attachments successfully")
        
        return results
    
    def create_attachment_document(self, issue_id: str, attachment: Dict[str, Any],
                                   owner: str = '', repo: str = '') -> Dict[str, Any]:
        """
        Create a document suitable for vector database indexing in issue_history.
        
        Args:
            issue_id: The issue ID this attachment belongs to
            attachment: Processed attachment data
            owner: Repository/project owner
            repo: Repository/project name
            
        Returns:
            Document ready for indexing in issue_history
        """
        content = attachment.get('content', '')
        
        # Truncate very long content
        max_content_length = 50000  # ~50KB of text
        if len(content) > max_content_length:
            content = content[:max_content_length] + "\n\n[Content truncated...]"
        
        # Create text preview
        text_preview = content[:500] + '...' if len(content) > 500 else content
        
        return {
            # Document identifiers - matching issue_history schema
            'document_id': f"{issue_id}_attachment_{attachment.get('filename', 'unknown')}",
            'document_type': 'attachment',
            'issue_id': issue_id,
            
            # Repository identifiers
            'repo_owner': owner,
            'repo_name': repo,
            'repo_full_name': f"{owner}/{repo}" if owner and repo else '',
            
            # Attachment-specific fields
            'filename': attachment.get('filename', 'unknown'),
            'file_type': attachment.get('file_type', 'unknown'),
            'content_type': attachment.get('mime_type', ''),
            'file_size': attachment.get('size', 0),
            'text_content': content,
            'text_length': len(content),
            'text_preview': text_preview,
            'attachment_url': attachment.get('url', ''),
            
            # Metadata
            'created_by': attachment.get('author', 'unknown'),
            'created_at': attachment.get('created', ''),
        }
